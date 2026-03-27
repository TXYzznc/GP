#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TXT to XLSX Converter
将txt格式的配置表转换为xlsx格式
"""

import os
import sys
import re
import subprocess
import platform
from openpyxl import Workbook
from openpyxl.styles import Alignment, numbers


def open_folder(folder_path):
    """跨平台打开文件夹"""
    try:
        folder_path = os.path.abspath(folder_path)
        
        system = platform.system()
        if system == 'Windows':
            # Windows: 使用 explorer
            subprocess.Popen(['explorer', folder_path])
        elif system == 'Darwin':
            # macOS: 使用 open
            subprocess.Popen(['open', folder_path])
        else:
            # Linux: 使用 xdg-open
            subprocess.Popen(['xdg-open', folder_path])
        
        print(f"✓ 已打开文件夹: {folder_path}")
        return True
    except Exception as e:
        print(f"⚠ 无法打开文件夹: {e}")
        return False


class DataTableProcessor:
    """数据表处理器 - 参考DataTableGenerator的创建方式"""
    
    def __init__(self, data_table_file, encoding='utf-8-sig', name_row=1, type_row=2, 
                 default_value_row=None, comment_row=3, content_start_row=4, id_column=1):
        """
        初始化数据表处理器
        参数对应DataTableGenerator.CreateDataTableProcessor的参数：
        - name_row: 1 (字段名在第2行，0-indexed)
        - type_row: 2 (字段类型在第3行，0-indexed)
        - default_value_row: None (无默认值行)
        - comment_row: 3 (注释在第4行，0-indexed)
        - content_start_row: 4 (数据从第5行开始，0-indexed)
        - id_column: 1 (ID列在第2列，0-indexed)
        """
        self.data_table_file = data_table_file
        self.encoding = encoding
        self.name_row = name_row
        self.type_row = type_row
        self.default_value_row = default_value_row
        self.comment_row = comment_row
        self.content_start_row = content_start_row
        self.id_column = id_column
        
        # 读取并解析文件
        self.raw_values = []
        self.load_raw_data()
        
        # 提取各行数据
        self.name_row_data = self.raw_values[self.name_row] if self.name_row < len(self.raw_values) else []
        self.type_row_data = self.raw_values[self.type_row] if self.type_row < len(self.raw_values) else []
        self.comment_row_data = self.raw_values[self.comment_row] if self.comment_row is not None and self.comment_row < len(self.raw_values) else []
        
        # 扩展的类型映射 - 基于DataTableGenerator支持的所有类型
        self.type_mapping = {
            # 基础数值类型
            'int': int,
            'int32': int,
            'system.int32': int,
            'uint': int,  # 无符号整数，Python中统一为int
            'system.uint32': int,
            'short': int,
            'system.int16': int,
            'ushort': int,
            'system.uint16': int,
            'long': int,
            'system.int64': int,
            'ulong': int,
            'system.uint64': int,
            'byte': int,
            'system.byte': int,
            'sbyte': int,
            'system.sbyte': int,
            
            # 浮点数类型
            'float': float,
            'system.single': float,
            'double': float,
            'system.double': float,
            'decimal': float,
            'system.decimal': float,
            
            # 字符串和字符类型
            'string': str,
            'system.string': str,
            'char': str,
            'system.char': str,
            
            # 布尔类型
            'bool': bool,
            'system.boolean': bool,
            
            # 日期时间类型
            'datetime': str,  # 存储为字符串，需要特殊解析
            'system.datetime': str,
            
            # Unity特有类型（存储为字符串，需要特殊解析）
            'vector2': str,
            'unityengine.vector2': str,
            'vector2int': str,
            'unityengine.vector2int': str,
            'vector3': str,
            'unityengine.vector3': str,
            'vector3int': str,
            'unityengine.vector3int': str,
            'vector4': str,
            'unityengine.vector4': str,
            'quaternion': str,
            'unityengine.quaternion': str,
            'color': str,
            'unityengine.color': str,
            'color32': str,
            'unityengine.color32': str,
            'rect': str,
            'unityengine.rect': str,
            'int4': str,  # Unity Mathematics类型
            
            # 枚举类型
            'enum': str,  # 枚举值存储为字符串
            'system.enum': str,
            
            # 数组类型
            'int[]': str,
            'int32[]': str,
            'system.int32[]': str,
            'uint[]': str,
            'float[]': str,
            'double[]': str,
            'bool[]': str,
            'string[]': str,
            'long[]': str,
            'vector2[]': str,
            'vector2int[]': str,
            'vector3[]': str,
            'vector3int[]': str,
            'vector4[]': str,
            'int4[]': str,
            
            # 二维数组类型
            'int[,]': str,
            'bool[,]': str,
            'double[,]': str,
            'float[,]': str,
            
            # 特殊类型
            'id': int,  # ID类型
            'comment': str,  # 注释类型
            'type': str,  # 类型字段
        }
    
    def load_raw_data(self):
        """加载原始数据，按DataTableProcessor的方式处理"""
        with open(self.data_table_file, 'r', encoding=self.encoding) as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.rstrip('\n\r')
            if not line.strip():  # 跳过空行
                continue
            
            # 按制表符分割并去除首尾空白
            raw_value = line.split('\t')
            raw_value = [cell.strip() for cell in raw_value]
            self.raw_values.append(raw_value)
    
    def get_raw_row_count(self):
        """获取原始行数"""
        return len(self.raw_values)
    
    def get_raw_column_count(self):
        """获取原始列数（返回第二行字段名行的列数）"""
        if not self.raw_values or len(self.raw_values) < 2:
            return 0
        return len(self.raw_values[1])  # 返回第二行（字段名行）的列数
    
    def get_name(self, raw_column):
        """获取指定列的字段名"""
        if raw_column < len(self.name_row_data):
            return self.name_row_data[raw_column]
        return ""
    
    def get_type(self, raw_column):
        """获取指定列的字段类型"""
        if raw_column < len(self.type_row_data):
            return self.type_row_data[raw_column]
        return "string"
    
    def get_comment(self, raw_column):
        """获取指定列的注释"""
        if self.comment_row_data and raw_column < len(self.comment_row_data):
            return self.comment_row_data[raw_column]
        return ""
    
    def get_value(self, raw_row, raw_column):
        """获取指定位置的值"""
        if raw_row < len(self.raw_values) and raw_column < len(self.raw_values[raw_row]):
            return self.raw_values[raw_row][raw_column]
        return ""
    
    def is_comment_row(self, raw_row):
        """判断是否为注释行"""
        if raw_row < len(self.raw_values):
            first_cell = self.raw_values[raw_row][0] if self.raw_values[raw_row] else ""
            return first_cell.startswith('#')
        return False
    
    def is_id_column(self, raw_column):
        """判断是否为ID列"""
        return raw_column == self.id_column
    
    def get_table_name(self):
        """获取表名（从第一行第二列）"""
        if len(self.raw_values) > 0 and len(self.raw_values[0]) > 1:
            return self.raw_values[0][1]
        return ""


class TxtToXlsxConverter:
    """TXT转XLSX转换器 - 使用DataTableProcessor方式"""
    
    def __init__(self, auto_open_folder=True):
        """
        初始化转换器
        
        参数:
            auto_open_folder: 是否在转换完成后自动打开输出文件夹（默认True）
        """
        self.auto_open_folder = auto_open_folder
        self._opened_folders = set()  # 记录已打开的文件夹，避免重复打开
    
    def create_data_table_processor(self, data_table_file):
        """创建数据表处理器 - 参考DataTableGenerator.CreateDataTableProcessor"""
        return DataTableProcessor(data_table_file, 'utf-8-sig', 1, 2, None, 3, 4, 1)
    
    def parse_txt_file_with_processor(self, txt_path):
        """使用DataTableProcessor方式解析txt文件"""
        processor = self.create_data_table_processor(txt_path)
        
        # 提取表结构信息
        table_name = processor.get_table_name()
        raw_column_count = processor.get_raw_column_count()
        
        # 构建字段信息
        field_row = []
        type_row = []
        comment_row = []
        
        for col in range(raw_column_count):
            field_row.append(processor.get_name(col))
            type_row.append(processor.get_type(col))
            comment_row.append(processor.get_comment(col))
        
        # 提取数据行
        data_rows = []
        for row in range(processor.content_start_row, processor.get_raw_row_count()):
            if processor.is_comment_row(row):
                continue
            
            data_row = []
            for col in range(raw_column_count):
                data_row.append(processor.get_value(row, col))
            data_rows.append(data_row)
        
        return {
            'table_name': table_name,
            'field_row': field_row,
            'type_row': type_row,
            'comment_row': comment_row,
            'data_rows': data_rows,
            'processor': processor
        }
    
    def convert_value_with_processor(self, value, field_type, processor=None):
        """使用DataTableProcessor方式转换值 - 支持所有DataTableGenerator类型"""
        if not value or value == '':
            return None
        
        # 统一转换为小写进行匹配
        field_type_lower = field_type.lower()
        
        try:
            # 基础数值类型
            if field_type_lower in ['int', 'int32', 'system.int32', 'uint', 'system.uint32', 
                                   'short', 'system.int16', 'ushort', 'system.uint16',
                                   'long', 'system.int64', 'ulong', 'system.uint64',
                                   'byte', 'system.byte', 'sbyte', 'system.sbyte', 'id']:
                return int(float(value))
            
            # 浮点数类型
            elif field_type_lower in ['float', 'system.single', 'double', 'system.double', 
                                     'decimal', 'system.decimal']:
                return float(value)
            
            # 布尔类型
            elif field_type_lower in ['bool', 'system.boolean']:
                return value.lower() in ['true', '1', 'yes']
            
            # 字符串和字符类型
            elif field_type_lower in ['string', 'system.string', 'char', 'system.char']:
                return str(value)
            
            # 日期时间类型 - 保持原始字符串格式
            elif field_type_lower in ['datetime', 'system.datetime']:
                return str(value)  # 可以添加日期验证
            
            # Unity向量类型 - 保持原始字符串格式
            elif field_type_lower in ['vector2', 'unityengine.vector2', 'vector2int', 'unityengine.vector2int',
                                     'vector3', 'unityengine.vector3', 'vector3int', 'unityengine.vector3int',
                                     'vector4', 'unityengine.vector4', 'quaternion', 'unityengine.quaternion']:
                return str(value)  # 格式如: "1.0,2.0,3.0"
            
            # Unity颜色类型 - 保持原始字符串格式
            elif field_type_lower in ['color', 'unityengine.color', 'color32', 'unityengine.color32']:
                return str(value)  # 格式如: "1.0,0.5,0.0,1.0"
            
            # Unity其他类型
            elif field_type_lower in ['rect', 'unityengine.rect', 'int4']:
                return str(value)
            
            # 枚举类型 - 保持原始字符串格式
            elif field_type_lower in ['enum', 'system.enum']:
                return str(value)  # 格式如: "EnumType.Value1"
            
            # 数组类型 - 保持原始字符串格式
            elif field_type_lower.endswith('[]') or field_type_lower.endswith('[,]'):
                return str(value)  # 格式如: "1,2,3,4" 或 "1,2|3,4"
            
            # 注释和类型字段
            elif field_type_lower in ['comment', 'type']:
                return str(value)
            
            # 默认处理为字符串
            else:
                return str(value)
                
        except Exception as e:
            # 转换失败时返回原始字符串
            print(f"警告: 类型转换失败 '{field_type}': {value} -> {e}")
            return str(value)
    
    def create_xlsx_with_processor(self, parsed_data, output_path):
        """使用DataTableProcessor方式创建xlsx文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        
        processor = parsed_data['processor']
        field_row = parsed_data['field_row']
        type_row = parsed_data['type_row']
        comment_row = parsed_data['comment_row']
        table_name = parsed_data['table_name']
        
        # 识别数组类型的列索引（用于设置文本格式）
        array_columns = set()
        for col_idx in range(1, len(type_row)):
            if col_idx < len(type_row):
                field_type = type_row[col_idx].strip().lower() if type_row[col_idx] else ''
                # 检查是否为数组类型（包含[]或[,]）
                if '[]' in field_type or '[,]' in field_type:
                    array_columns.add(col_idx + 1)  # +1 因为Excel列从1开始
        
        # 按DataTableProcessor的行结构创建xlsx
        # 使用逐单元格写入方式，确保所有列（包括空列）都被正确写入
        
        # 第1行 (索引0): # + 表名 + 其余为None
        ws.cell(row=1, column=1, value='#')
        ws.cell(row=1, column=2, value=table_name)
        for col_idx in range(3, len(field_row) + 1):
            ws.cell(row=1, column=col_idx, value=None)
        
        # 第2行 (索引1): # + 字段名
        ws.cell(row=2, column=1, value='#')
        for col_idx, field_name in enumerate(field_row[1:], start=2):
            ws.cell(row=2, column=col_idx, value=field_name if field_name else None)
        
        # 第3行 (索引2): # + 字段类型
        ws.cell(row=3, column=1, value='#')
        for col_idx, field_type in enumerate(type_row[1:], start=2):
            ws.cell(row=3, column=col_idx, value=field_type if field_type else None)
        
        # 第4行 (索引3): # + 注释
        ws.cell(row=4, column=1, value='#')
        if comment_row:
            for col_idx, comment in enumerate(comment_row[1:], start=2):
                ws.cell(row=4, column=col_idx, value=comment if comment else None)
        else:
            for col_idx in range(2, len(field_row) + 1):
                ws.cell(row=4, column=col_idx, value=None)
        
        # 数据行 (从索引4开始)
        for row_idx, data_row in enumerate(parsed_data['data_rows'], start=5):
            # 第一列为空（None）
            ws.cell(row=row_idx, column=1, value=None)
            
            # 从第二列开始处理数据
            for col_idx in range(1, len(field_row)):
                if col_idx < len(data_row):
                    value = data_row[col_idx].strip() if data_row[col_idx] else ''
                    
                    # 如果这一列是空的，直接设为None
                    if not value:
                        ws.cell(row=row_idx, column=col_idx + 1, value=None)
                    else:
                        # 根据字段类型转换值
                        if col_idx < len(type_row):
                            field_type = type_row[col_idx].strip() if type_row[col_idx] else 'string'
                            converted_value = self.convert_value_with_processor(value, field_type, processor)
                            cell = ws.cell(row=row_idx, column=col_idx + 1, value=converted_value)
                            
                            # 如果是数组类型列，设置为文本格式（避免Excel自动格式化）
                            if (col_idx + 1) in array_columns:
                                cell.number_format = numbers.FORMAT_TEXT
                        else:
                            ws.cell(row=row_idx, column=col_idx + 1, value=value)
                else:
                    ws.cell(row=row_idx, column=col_idx + 1, value=None)
        
        # 设置单元格对齐方式和数组列的文本格式
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # 为数组列的所有单元格设置文本格式
                if cell.column in array_columns:
                    cell.number_format = numbers.FORMAT_TEXT
        
        # 设置数据行（从第5行开始）的行高为固定值（15磅，约1行高度）
        for row_idx in range(5, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15
        
        # 自动调整列宽（根据内容）
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # 计算单元格内容的长度（考虑中文字符占2个宽度）
                        cell_value = str(cell.value)
                        # 简单估算：中文字符算2个宽度，英文算1个
                        length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            # 设置列宽（加一些边距，最小宽度8，最大宽度50）
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(output_path)
        print(f"✓ 已生成: {output_path}")
        if array_columns:
            array_col_names = [field_row[col-1] for col in array_columns if col-1 < len(field_row)]
            print(f"  已将数组列设置为文本格式: {', '.join(array_col_names)}")
        print(f"  已设置列宽自适应和数据行固定行高")
    
    def convert_file_with_processor(self, txt_path, output_dir=None):
        """使用DataTableProcessor方式转换单个txt文件为xlsx"""
        if not os.path.exists(txt_path):
            print(f"✗ 文件不存在: {txt_path}")
            return False
        
        # 使用DataTableProcessor方式解析txt文件
        try:
            parsed_data = self.parse_txt_file_with_processor(txt_path)
        except Exception as e:
            print(f"✗ 解析文件失败 {txt_path}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        # 确定输出路径
        if output_dir is None:
            output_dir = "output"
        
        os.makedirs(output_dir, exist_ok=True)
        
        # 生成输出文件名
        base_name = os.path.basename(txt_path)
        xlsx_name = os.path.splitext(base_name)[0] + '.xlsx'
        output_path = os.path.join(output_dir, xlsx_name)
        
        # 创建xlsx文件
        try:
            self.create_xlsx_with_processor(parsed_data, output_path)
            
            # 转换成功后，自动打开输出文件夹
            if self.auto_open_folder:
                abs_output_dir = os.path.abspath(output_dir)
                if abs_output_dir not in self._opened_folders:
                    open_folder(abs_output_dir)
                    self._opened_folders.add(abs_output_dir)
            
            return True
        except Exception as e:
            print(f"✗ 生成xlsx失败 {output_path}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def convert_directory_with_processor(self, input_dir, output_dir=None):
        """使用DataTableProcessor方式转换整个目录的txt文件"""
        if not os.path.exists(input_dir):
            print(f"✗ 目录不存在: {input_dir}")
            return
        
        txt_files = [f for f in os.listdir(input_dir) if f.endswith('.txt')]
        
        if not txt_files:
            print(f"✗ 目录中没有找到txt文件: {input_dir}")
            return
        
        print(f"找到 {len(txt_files)} 个txt文件")
        success_count = 0
        
        for txt_file in txt_files:
            txt_path = os.path.join(input_dir, txt_file)
            if self.convert_file_with_processor(txt_path, output_dir):
                success_count += 1
        
        print(f"\n转换完成: {success_count}/{len(txt_files)} 个文件成功")
        
        # 批量转换完成后，打开输出文件夹
        if self.auto_open_folder and success_count > 0:
            if output_dir is None:
                output_dir = "output"
            abs_output_dir = os.path.abspath(output_dir)
            if abs_output_dir not in self._opened_folders:
                open_folder(abs_output_dir)
                self._opened_folders.add(abs_output_dir)
    
    # 保留原有方法以保持向后兼容
    def parse_txt_file(self, txt_path):
        """解析txt文件，提取表结构和数据（原有方法，保持兼容性）"""
        return self.parse_txt_file_with_processor(txt_path)
    
    def convert_value(self, value, field_type):
        """根据字段类型转换值（原有方法，保持兼容性）"""
        return self.convert_value_with_processor(value, field_type)
    
    def create_xlsx(self, parsed_data, output_path):
        """根据解析的数据创建xlsx文件（原有方法，保持兼容性）"""
        return self.create_xlsx_with_processor(parsed_data, output_path)
    
    def convert_file(self, txt_path, output_dir=None):
        """转换单个txt文件为xlsx（原有方法，保持兼容性）"""
        return self.convert_file_with_processor(txt_path, output_dir)
    
    def convert_directory(self, input_dir, output_dir=None):
        """转换整个目录的txt文件（原有方法，保持兼容性）"""
        return self.convert_directory_with_processor(input_dir, output_dir)


def main():
    """主函数"""
    converter = TxtToXlsxConverter()
    
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  转换单个文件: python txt_to_xlsx_converter.py <txt文件路径> [输出目录]")
        print("  转换整个目录: python txt_to_xlsx_converter.py <txt目录路径> [输出目录]")
        print("\n示例:")
        print("  python txt_to_xlsx_converter.py BuffTable.txt")
        print("  python txt_to_xlsx_converter.py BuffTable.txt output")
        print('  python txt_to_xlsx_converter.py "txt格式源文件（规定字段和内容）" output')
        return
    
    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "output"
    
    if os.path.isfile(input_path):
        converter.convert_file(input_path, output_dir)
    elif os.path.isdir(input_path):
        converter.convert_directory(input_path, output_dir)
    else:
        print(f"✗ 路径不存在: {input_path}")


if __name__ == "__main__":
    main()

from openpyxl import load_workbook

wb = load_workbook('D:/Sourcetree/Clash_Of_Gods/AAAGameData/DataTables/SummonerSkillTable.xlsx')
ws = wb.active

print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

print("\n表头行（第1行）:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(1, col)
    print(f"Col {col}: {cell.value}")

print("\n第5行（ID=101）:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(5, col)
    print(f"Col {col}: {cell.value}")

print("\n第6行（ID=102）:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(6, col)
    print(f"Col {col}: {cell.value}")

print("\n第7行（ID=103）:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(7, col)
    print(f"Col {col}: {cell.value}")

print("\n第8行（ID=104）:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(8, col)
    print(f"Col {col}: {cell.value}")

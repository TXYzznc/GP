#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从原始TXT文件直接读取数据，生成正确的XLSX文件
"""

import openpyxl

# 读取原始TXT文件
with open('Assets/AAAGame/DataTable/SummonerSkillTable.txt', 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()

# 创建新的工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet 1"

# 写入前4行（注释行和列头）
for row_idx in range(4):
    line = lines[row_idx].rstrip('\n\r')
    cells = line.split('\t')
    for col_idx, cell_value in enumerate(cells, 1):
        ws.cell(row_idx + 1, col_idx, cell_value)

# 写入数据行（从第5行开始）
for row_idx in range(4, len(lines)):
    line = lines[row_idx].rstrip('\n\r')
    if not line or line.startswith('#'):
        continue
    
    cells = line.split('\t')
    for col_idx, cell_value in enumerate(cells, 1):
        ws.cell(row_idx - 3, col_idx, cell_value)

# 设置列宽
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 15
for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
    ws.column_dimensions[col].width = 12

# 保存文件
wb.save('AAAGameData/DataTables/SummonerSkillTable.xlsx')
print("✓ 已从TXT文件直接生成XLSX文件")
print("✓ 输出文件: AAAGameData/DataTables/SummonerSkillTable.xlsx")

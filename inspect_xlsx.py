#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查XLSX文件的结构
"""

import openpyxl

# 打开XLSX文件
wb = openpyxl.load_workbook('AAAGameData/DataTables/SummonerSkillTable.xlsx')
ws = wb.active

print('=== XLSX文件结构检查 ===\n')
print(f'工作表名称: {ws.title}')
print(f'最大行数: {ws.max_row}')
print(f'最大列数: {ws.max_column}')
print()

# 打印前10行
print('前10行数据:')
for row_idx in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(6, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        row_data.append(str(cell.value)[:20] if cell.value else "")
    print(f'行 {row_idx}: {row_data}')

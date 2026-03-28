from openpyxl import load_workbook

# 读取召唤师技能表
wb_skills = load_workbook('AAAGameData/DataTables/SummonerSkillTable.xlsx')
ws_skills = wb_skills.active

print('=== SummonerSkillTable ===')
for i, row in enumerate(ws_skills.iter_rows(values_only=True)):
    if i < 10:  # 只打印前10行
        print(f"Row {i}: {row}")
    if i >= 4 and row[0] is not None and row[0] != '#':  # 数据行
        print(f"\n技能: {row[3]} (ID={row[1]})")
        print(f"描述: {row[-1]}")

print('\n\n=== BuffTable ===')
wb_buff = load_workbook('AAAGameData/DataTables/BuffTable.xlsx')
ws_buff = wb_buff.active

for i, row in enumerate(ws_buff.iter_rows(values_only=True)):
    if i < 10:
        print(f"Row {i}: {row}")
    if i >= 4 and row[0] is not None and row[0] != '#':
        if row[1] in [4001, 4002]:
            print(f"\nBuff: {row[3]} (ID={row[1]})")
            print(f"描述: {row[-1]}")

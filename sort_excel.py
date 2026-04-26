from openpyxl import load_workbook

file_path = 'AI工作区/FAULTY METER LIST ABA.xlsx'
wb = load_workbook(file_path)
sheet = wb.active

# 读取所有数据（包括表头）
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

print("原始数据:")
for row in data[:10]:
    print(row)

# 分离表头和数据
header = data[0]
rows = data[1:]

print(f"\n列数: {len(header)}")
print(f"第二列索引: 1 (列 {chr(65+1)})")

# 按第二列（索引1）升序排序 - 转换为数字排序
sorted_rows = sorted(rows, key=lambda x: int(x[1]) if x[1] and str(x[1]).isdigit() else 0)

# 创建新工作簿
from openpyxl import Workbook
new_wb = Workbook()
new_sheet = new_wb.active

# 写入表头
for col_idx, value in enumerate(header, 1):
    new_sheet.cell(row=1, column=col_idx, value=value)

# 写入排序后的数据
for row_idx, row_data in enumerate(sorted_rows, 2):
    for col_idx, value in enumerate(row_data, 1):
        new_sheet.cell(row=row_idx, column=col_idx, value=value)

# 保存
output_path = 'AI工作区/FAULTY METER LIST ABA_sorted.xlsx'
new_wb.save(output_path)

print(f"\n排序完成！已保存到: {output_path}")
print("\n排序后的前10行:")
for i, row in enumerate(sorted_rows[:10], 1):
    print(f"{i}: {row}")

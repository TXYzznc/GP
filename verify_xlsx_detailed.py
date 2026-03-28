#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
详细验证XLSX文件中的数据
"""

import openpyxl

# 打开XLSX文件
wb = openpyxl.load_workbook('AAAGameData/DataTables/SummonerSkillTable.xlsx')
ws = wb.active

print('=== 详细验证XLSX文件 ===\n')

# 预期的数据
expected_data = {
    101: {'name': '狂怒之心', 'params': '0.5'},
    102: {'name': '战意激昂', 'params': '20'},
    103: {'name': '王者号令', 'params': '0.5'},
    104: {'name': '钢铁洪流', 'params': '4'},
    105: {'name': '天灾降临', 'params': '2.0,1.5,0.3'},
    106: {'name': '寂灭斩', 'params': '2.5'},
    107: {'name': '孤影', 'params': '0.4,0.6'},
    108: {'name': '裁决之刻', 'params': '1.8,0.7'},
}

# 遍历数据行
for row_idx in range(4, ws.max_row + 1):
    skill_id_cell = ws.cell(row_idx, 2)
    skill_name_cell = ws.cell(row_idx, 3)
    params_cell = ws.cell(row_idx, 22)
    
    if skill_id_cell.value is None:
        break
    
    try:
        skill_id = int(skill_id_cell.value)
    except:
        continue
    
    if skill_id in expected_data:
        expected = expected_data[skill_id]
        actual_name = skill_name_cell.value
        actual_params = params_cell.value
        
        name_match = actual_name == expected['name']
        params_match = str(actual_params) == str(expected['params'])
        
        status = '✓' if (name_match and params_match) else '❌'
        print(f'{status} ID {skill_id}:')
        print(f'   Name: {actual_name} (期望: {expected["name"]})')
        print(f'   Params: {actual_params} (期望: {expected["params"]})')
        print()

print('✓ 验证完成')

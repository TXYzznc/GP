#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证XLSX文件中的Params列数据是否正确
"""

import openpyxl

# 打开XLSX文件
wb = openpyxl.load_workbook('AAAGameData/DataTables/SummonerSkillTable.xlsx')
ws = wb.active

print('=== 验证XLSX文件中的Params列 ===\n')

# 预期的Params值（根据TXT文件）
expected_params = {
    101: '0',
    102: '20',
    103: '0.5',
    104: '4',
    105: '2.0,1.5,0.3',
    106: '2.5',
    107: '0.4,0.6',
    108: '1.8,0.7',
    201: '0.1,0.8,0.3',
    202: '0,0.20,0.15',
}

# 遍历数据行（从第4行开始，跳过注释行）
for row_idx in range(4, ws.max_row + 1):
    skill_id = ws.cell(row_idx, 2).value  # ID列
    skill_name = ws.cell(row_idx, 3).value  # Name列
    params_value = ws.cell(row_idx, 22).value  # Params列（第22列）
    
    if skill_id is None:
        break
    
    skill_id = int(skill_id) if isinstance(skill_id, (int, float)) else skill_id
    
    if skill_id in expected_params:
        expected = expected_params[skill_id]
        status = '✓' if str(params_value) == str(expected) else '❌'
        print(f'{status} ID {skill_id} ({skill_name}): Params = {params_value}')
        if str(params_value) != str(expected):
            print(f'   期望值: {expected}')

print('\n✓ 验证完成')

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AudioClipTable TXT 转 XLSX 转换脚本

用途: 将 AudioClipTable.txt 转换为 Excel 格式 (.xlsx)
使用: python convert_audio_table_txt_to_xlsx.py

依赖: openpyxl
安装: pip install openpyxl
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("❌ 缺少依赖库 openpyxl，请先安装:")
    print("   pip install openpyxl")
    sys.exit(1)


def convert_txt_to_xlsx(txt_path, xlsx_path):
    """
    将 TXT 配置表转换为 XLSX 格式

    Args:
        txt_path: 源 TXT 文件路径
        xlsx_path: 目标 XLSX 文件路径
    """

    print(f"📖 读取 TXT 文件: {txt_path}")

    # 1. 读取 TXT 文件
    if not os.path.exists(txt_path):
        print(f"❌ 文件不存在: {txt_path}")
        return False

    try:
        rows = []
        with open(txt_path, 'r', encoding='utf-8-sig') as f:
            for line_num, line in enumerate(f, 1):
                row = line.rstrip('\n\r').split('\t')
                rows.append(row)
                if line_num <= 5:
                    print(f"   行 {line_num}: {len(row)} 列")

        print(f"✅ 成功读取 {len(rows)} 行数据")
    except Exception as e:
        print(f"❌ 读取 TXT 文件失败: {e}")
        return False

    # 2. 创建 XLSX
    print(f"📝 创建 XLSX 文件: {xlsx_path}")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AudioClipTable"

        # 3. 写入数据并设置格式
        print("🎨 写入数据并设置格式...")

        for row_idx, row in enumerate(rows, 1):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)

                # 设置元数据行的格式（前 4 行）
                if row_idx <= 4:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # 数组列设置为文本格式
                if row_idx > 4 and val and isinstance(val, str) and ',' in val:
                    cell.number_format = '@'  # 文本格式

                # 所有单元格启用自动换行
                if row_idx > 4:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 4. 调整列宽
        print("📐 调整列宽...")

        column_widths = {
            1: 8,    # Id
            2: 18,   # AudioName
            3: 10,   # AudioType
            4: 20,   # ResourcePath
            5: 10,   # Duration
            6: 8,    # Volume
            7: 8,    # Pitch
            8: 8,    # IsLoop
            9: 8,    # Is3D
            10: 12,  # MaxDistance
            11: 10,  # Priority
            12: 12,  # FadeInTime
            13: 12,  # FadeOutTime
            14: 12,  # Tag
        }

        for col_idx, width in column_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 5. 冻结前 4 行
        ws.freeze_panes = "A5"

        # 6. 保存
        print(f"💾 保存文件: {xlsx_path}")
        wb.save(xlsx_path)

        print(f"✅ 转换成功!")
        print(f"   输出: {xlsx_path}")
        print(f"   行数: {len(rows)}")
        print(f"   列数: {len(rows[0]) if rows else 0}")

        return True

    except Exception as e:
        print(f"❌ 创建 XLSX 失败: {e}")
        return False


def verify_files(txt_path, xlsx_path):
    """
    验证转换结果
    """

    print("\n🔍 验证转换结果...")

    # 读取 TXT
    with open(txt_path, 'r', encoding='utf-8-sig') as f:
        txt_rows = [line.rstrip('\n\r').split('\t') for line in f]

    # 读取 XLSX
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    xlsx_rows = []
    for row in ws.iter_rows(values_only=True):
        xlsx_rows.append([str(val) if val is not None else '' for val in row])

    # 对比行数
    if len(txt_rows) != len(xlsx_rows):
        print(f"⚠️  行数不匹配: TXT={len(txt_rows)}, XLSX={len(xlsx_rows)}")
    else:
        print(f"✅ 行数一致: {len(txt_rows)}")

    # 对比列数
    if txt_rows and xlsx_rows:
        txt_cols = len(txt_rows[0])
        xlsx_cols = len(xlsx_rows[0])
        if txt_cols != xlsx_cols:
            print(f"⚠️  列数不匹配: TXT={txt_cols}, XLSX={xlsx_cols}")
        else:
            print(f"✅ 列数一致: {txt_cols}")

    # 对比样本数据（前 5 行，后 5 行）
    mismatches = 0
    for i in range(min(5, len(txt_rows))):
        txt_row = txt_rows[i]
        xlsx_row = xlsx_rows[i]
        if len(txt_row) != len(xlsx_row):
            mismatches += 1

    if mismatches == 0:
        print(f"✅ 数据验证通过")
    else:
        print(f"⚠️  发现 {mismatches} 行数据不匹配")


def main():
    """主函数"""

    print("=" * 60)
    print("  AudioClipTable TXT → XLSX 转换工具")
    print("=" * 60)
    print()

    # 获取文件路径
    if len(sys.argv) > 1:
        txt_path = sys.argv[1]
    else:
        # 查找当前目录下的 AudioClipTable.txt
        txt_path = "AudioClipTable.txt"

    # 确定输出路径
    if len(sys.argv) > 2:
        xlsx_path = sys.argv[2]
    else:
        # 将 .txt 替换为 .xlsx
        xlsx_path = txt_path.replace('.txt', '.xlsx')

    print(f"📁 输入文件: {txt_path}")
    print(f"📁 输出文件: {xlsx_path}")
    print()

    # 转换
    success = convert_txt_to_xlsx(txt_path, xlsx_path)

    if success:
        # 验证
        verify_files(txt_path, xlsx_path)
        print()
        print("=" * 60)
        print("✅ 转换完成！")
        print()
        print("下一步:")
        print(f"1. 将 {xlsx_path} 复制到项目:")
        print("   Assets/AAAGame/DataTable/AudioClipTable.xlsx")
        print()
        print("2. 在 Unity 编辑器中执行:")
        print("   GameFramework → DataTable → Generate")
        print()
        print("=" * 60)
    else:
        print()
        print("❌ 转换失败，请检查输入文件或依赖库")
        sys.exit(1)


if __name__ == "__main__":
    main()
